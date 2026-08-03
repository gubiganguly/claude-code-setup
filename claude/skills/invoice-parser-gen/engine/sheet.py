"""Write the extraction workbook that lands in the invoice folder.

Three sheets: Invoices (one row each), Line Items (one row per charge, with the
invoice fields denormalised on so every row stands alone for pivoting), and
Exceptions (only what needs a human).

Two rules shape the implementation.

Rows that did not fully verify are filled red, and verified rows are left plain.
Colouring both would make the sheet a wash of pastel and bury the three rows
that matter; a single colour with one meaning stays scannable at 500 rows.

Writes are append-only. A re-run never rewrites an existing row, which is what
lets the trailing Notes column survive: anything typed there is still there
after the next run. Rebuilding the sheet from scratch each time would be simpler
and would silently eat the user's annotations.
"""

from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

__all__ = ["ExtractionWorkbook", "DEFAULT_FILENAME"]

DEFAULT_FILENAME = "Invoice Extraction.xlsx"

SHEET_INVOICES = "Invoices"
SHEET_LINE_ITEMS = "Line Items"
SHEET_EXCEPTIONS = "Exceptions"

# Warm neutral header, and a red reserved exclusively for "not fully verified".
_HEADER_FILL = PatternFill("solid", start_color="FF2A2622", end_color="FF2A2622")
_HEADER_FONT = Font(color="FFF7F5F2", bold=True, size=10)
_FLAG_FILL = PatternFill("solid", start_color="FFFCE1DE", end_color="FFFCE1DE")
_FLAG_FONT = Font(color="FF8C1D18", bold=True, size=10)
_MONEY_FORMAT = "#,##0.00"

INVOICE_COLUMNS = [
    ("Invoice #", 14),
    ("Invoice Date", 12),
    ("Status", 20),
    ("Pages", 7),
    ("Line Items", 10),
    ("Computed Total", 15),
    ("Printed Subtotal", 16),
    ("Tax", 11),
    ("Printed Total", 14),
    ("Delta", 10),
    ("Checks", 9),
    ("Depth", 7),
    ("Departments", 26),
    ("Service Ticket", 15),
    ("Customer Ref", 14),
    ("User ID", 20),
    ("Sold To", 11),
    ("Payer", 11),
    ("Payment Terms", 14),
    ("Sort #", 14),
    ("Route", 22),
    ("OCR", 6),
    ("Source File", 20),
    ("Extracted", 20),
    ("Notes", 34),
]

LINE_ITEM_COLUMNS = [
    ("Invoice #", 14),
    ("Invoice Date", 12),
    ("Department", 20),
    ("Employee", 11),
    ("Material", 12),
    ("Description", 46),
    ("Variant", 9),
    ("Freq", 6),
    ("Exch", 6),
    ("Qty", 8),
    ("Unit Price", 11),
    ("Line Total", 12),
    ("Taxable", 8),
    ("Page", 6),
]

EXCEPTION_COLUMNS = [
    ("Invoice #", 14),
    ("Status", 22),
    ("Reason", 96),
    ("Source File", 20),
    ("Extracted", 20),
]

_MONEY_COLUMNS = {
    SHEET_INVOICES: {"Computed Total", "Printed Subtotal", "Tax", "Printed Total", "Delta"},
    SHEET_LINE_ITEMS: {"Unit Price", "Line Total"},
    SHEET_EXCEPTIONS: set(),
}


def _as_number(value):
    """Money reaches the cell as a number so the user can sum a column."""
    if isinstance(value, Decimal):
        return float(value)
    return value


class ExtractionWorkbook:
    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)
        self._book: Workbook | None = None

    # -- open / create -----------------------------------------------------

    @property
    def book(self) -> Workbook:
        if self._book is None:
            self._book = (
                load_workbook(self.path) if self.path.is_file() else self._create()
            )
        return self._book

    def _create(self) -> Workbook:
        book = Workbook()
        book.remove(book.active)
        for name, columns in (
            (SHEET_INVOICES, INVOICE_COLUMNS),
            (SHEET_LINE_ITEMS, LINE_ITEM_COLUMNS),
            (SHEET_EXCEPTIONS, EXCEPTION_COLUMNS),
        ):
            sheet = book.create_sheet(name)
            self._write_header(sheet, columns)
        return book

    @staticmethod
    def _write_header(sheet: Worksheet, columns: list[tuple[str, int]]) -> None:
        for index, (title, width) in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=index, value=title)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(columns))}1"
        )

    # -- read back ---------------------------------------------------------

    def existing_identities(self, column: str = "Invoice #") -> set[str]:
        """Invoice numbers already reported, read from the sheet itself.

        The sheet is what the user sees, so it is the authority on what has
        already been reported. Reading identities back from it means a deleted
        sidecar store cannot cause duplicate rows.
        """
        if not self.path.is_file():
            return set()
        sheet = self.book[SHEET_INVOICES]
        headers = [c.value for c in sheet[1]]
        if column not in headers:
            return set()
        at = headers.index(column) + 1
        found = set()
        for row in range(2, sheet.max_row + 1):
            value = sheet.cell(row=row, column=at).value
            if value not in (None, ""):
                found.add(str(value))
        return found

    # -- append ------------------------------------------------------------

    def _append(
        self, sheet_name: str, columns: list[tuple[str, int]], values: list[dict], *, flag: bool = False
    ) -> None:
        sheet = self.book[sheet_name]
        money = _MONEY_COLUMNS.get(sheet_name, set())
        titles = [title for title, _ in columns]
        for record in values:
            row = sheet.max_row + 1
            for index, title in enumerate(titles, start=1):
                cell = sheet.cell(row=row, column=index, value=_as_number(record.get(title)))
                if title in money:
                    cell.number_format = _MONEY_FORMAT
                    cell.alignment = Alignment(horizontal="right")
                if flag:
                    cell.fill = _FLAG_FILL
                    if index == 1 or titles[index - 1] == "Status":
                        cell.font = _FLAG_FONT

    def add_invoice(self, record: dict, line_items: list[dict]) -> None:
        flag = not record.get("_verified", False)
        self._append(SHEET_INVOICES, INVOICE_COLUMNS, [record], flag=flag)
        if line_items:
            self._append(SHEET_LINE_ITEMS, LINE_ITEM_COLUMNS, line_items)

    def add_exceptions(self, records: list[dict]) -> None:
        self._append(SHEET_EXCEPTIONS, EXCEPTION_COLUMNS, records, flag=True)

    def save(self) -> None:
        self.book.save(self.path)


def timestamp() -> str:
    return datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M")
